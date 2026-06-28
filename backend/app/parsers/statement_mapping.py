import os
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.normalization import normalize_label


DEFAULT_MAPPING_FILE = "PL1_TT201_2014_BTC.xlsx"
CODE_PATTERN = re.compile(r"^\d{2,3}[a-zA-Z]?$")


@dataclass(frozen=True)
class StatementMappingItem:
    statement_key: str
    code: str
    label: str
    raw_label: str
    normalized_label: str
    parent_code: str | None
    level: int
    row_number: int
    source_file: str


class StatementMapping:
    def __init__(self, items: list[StatementMappingItem], source_path: Path | None) -> None:
        self.items = items
        self.source_path = source_path
        self._by_statement_code: dict[tuple[str, str], StatementMappingItem] = {
            (item.statement_key, self.code_key(item.code)): item for item in items
        }

    @staticmethod
    def code_key(code: Any) -> str:
        return str(code or "").strip().lower()

    def get(self, statement_key: str, code: Any) -> StatementMappingItem | None:
        return self._by_statement_code.get((statement_key, self.code_key(code)))

    def find_by_label(self, statement_key: str, label: Any) -> StatementMappingItem | None:
        normalized = normalize_label(label)
        if not normalized:
            return None
        candidates = [
            item
            for item in self.items
            if item.statement_key == statement_key
            and (item.normalized_label in normalized or normalized in item.normalized_label)
        ]
        if not candidates:
            return None
        return max(candidates, key=lambda item: len(item.normalized_label))

    @property
    def source_name(self) -> str | None:
        return self.source_path.name if self.source_path else None


@lru_cache(maxsize=4)
def load_statement_mapping(path: str | None = None) -> StatementMapping:
    source_path = _resolve_mapping_path(path)
    if source_path is None or not source_path.exists():
        return StatementMapping([], source_path)

    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook["Noi dung Word"] if "Noi dung Word" in workbook.sheetnames else workbook.worksheets[0]
    items: list[StatementMappingItem] = []
    current_statement: str | None = None
    stack_by_statement: dict[str, dict[int, str]] = {}
    code_column_by_statement: dict[str, int] = {}

    for row_number in range(1, worksheet.max_row + 1):
        row_text = _row_text(worksheet, row_number)
        normalized_row = normalize_label(row_text)
        current_statement = _detect_statement_key(normalized_row, current_statement)
        if current_statement is not None:
            code_column = _find_header_code_column(worksheet, row_number)
            if code_column is not None:
                code_column_by_statement[current_statement] = code_column

        code_column = code_column_by_statement.get(current_statement or "", 3)
        code = _clean_code(worksheet.cell(row_number, code_column).value)
        raw_label = _clean_text(worksheet.cell(row_number, 1).value)
        if current_statement is None or code is None or not raw_label:
            continue

        label = _clean_template_label(raw_label)
        if not label:
            continue

        level = _infer_level(raw_label, code, current_statement)
        statement_stack = stack_by_statement.setdefault(current_statement, {})
        parent_code = _find_parent_code(statement_stack, level)
        statement_stack[level] = code
        for stale_level in [key for key in statement_stack if key > level]:
            statement_stack.pop(stale_level, None)

        items.append(
            StatementMappingItem(
                statement_key=current_statement,
                code=code,
                label=label,
                raw_label=raw_label,
                normalized_label=normalize_label(label),
                parent_code=parent_code,
                level=level,
                row_number=row_number,
                source_file=source_path.name,
            )
        )

    return StatementMapping(items, source_path)


def get_statement_mapping() -> StatementMapping:
    return load_statement_mapping(os.getenv("FINANCIAL_TEMPLATE_MAPPING_PATH"))


def _resolve_mapping_path(path: str | None) -> Path | None:
    candidates: list[Path] = []
    if path:
        candidates.append(Path(path))

    project_root = Path(__file__).resolve().parents[2]
    candidates.extend(
        [
            project_root / "test_data" / DEFAULT_MAPPING_FILE,
            Path("/app/test_data") / DEFAULT_MAPPING_FILE,
        ]
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0] if candidates else None


def _detect_statement_key(normalized_row: str, current_statement: str | None) -> str | None:
    if "bao cao ket qua hoat dong kinh doanh hop nhat" in normalized_row:
        return "income_statement"
    if "bao cao luu chuyen tien te hop nhat" in normalized_row:
        return "cash_flow"
    if "phuong phap truc tiep" in normalized_row:
        return "cash_flow_direct"
    if "phuong phap gian tiep" in normalized_row:
        return "cash_flow_indirect"
    if "bang can doi ke toan hop nhat" in normalized_row:
        return "financial_position"
    return current_statement


def _row_text(worksheet: Any, row_number: int) -> str:
    return " ".join(
        _clean_text(worksheet.cell(row_number, column).value)
        for column in range(1, worksheet.max_column + 1)
        if _clean_text(worksheet.cell(row_number, column).value)
    )


def _find_header_code_column(worksheet: Any, row_number: int) -> int | None:
    for column in range(1, worksheet.max_column + 1):
        if normalize_label(worksheet.cell(row_number, column).value) == "ma so":
            return column
    return None


def _clean_code(value: Any) -> str | None:
    text = _clean_text(value)
    if not CODE_PATTERN.fullmatch(text):
        return None
    return text


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def _clean_template_label(label: str) -> str:
    cleaned = _clean_text(label)
    cleaned = re.sub(r"^[A-Z]\s*[-–—]\s*", "", cleaned)
    cleaned = re.sub(r"^(?:[IVX]+|\d+)\s*[.)]\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"^-\s*", "", cleaned)
    return cleaned.strip(" .:-–—")


def _infer_level(raw_label: str, code: str, statement_key: str) -> int:
    label = raw_label.strip()
    if statement_key == "financial_position":
        if re.match(r"^[A-Z]\s*[-–—]", label):
            return 1
        if re.match(r"^(?:[IVX]+)\s*[.)]", label, flags=re.IGNORECASE):
            return 2
        if re.match(r"^(?:\d+|[-–—])\s*[.)]?", label):
            return 3
        return 1 if code.endswith("00") else 2 if code.endswith("0") else 3

    if statement_key == "income_statement":
        return 1

    if statement_key in {"cash_flow_direct", "cash_flow_indirect", "cash_flow"}:
        if re.match(r"^[-–—]", label):
            return 3
        if re.match(r"^\d+\s*[.)]", label):
            return 2
        return 1

    return 1


def _find_parent_code(stack: dict[int, str], level: int) -> str | None:
    parent_levels = [candidate_level for candidate_level in stack if candidate_level < level]
    if not parent_levels:
        return None
    return stack[max(parent_levels)]
